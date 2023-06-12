import openpyxl

def pindahkan_data_acc(nama_file):
    try:
        # Buka file Excel
        workbook = openpyxl.load_workbook(nama_file)
        
        # Mengambil sheet pertama
        sheet = workbook.active
        
        # Mencari data "acc" dalam kolom tertentu
        kolom_sumber = "G"  # Kolom yang ingin dicek
        baris_mulai = 2  # Nomor baris mulai
        
        data_acc = []
        for row in sheet.iter_rows(min_row=baris_mulai, values_only=True):
            nilai = row[0]  # Mengambil nilai kolom tertentu
            if nilai == "acc" or "Acc" or "ACC":
                data_acc.append(row)
        
        # Membuat file spreadsheet baru dan menyimpan data "acc" di dalamnya
        if data_acc:
            workbook_acc = openpyxl.Workbook()
            sheet_acc = workbook_acc.active
            
            for data in data_acc:
                sheet_acc.append(data)
            
            nama_file_acc = "baris_FT.xlsx"
            workbook_acc.save(nama_file_acc)
            print(f"Data dengan nilai 'acc' telah dipindahkan ke file '{nama_file_acc}'.")
        else:
            print("Tidak ada data dengan nilai 'acc' dalam kolom yang diperiksa.")
    
    except FileNotFoundError:
        print("File tidak ditemukan.")
    
    except Exception as e:
        print(f"Terjadi kesalahan: {str(e)}")

# Contoh penggunaan
nama_file = "hasil_pengajuan_tandatangan.xlsx"
pindahkan_data_acc(nama_file)