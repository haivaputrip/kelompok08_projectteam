def login():
    username = "admin"
    password = "password"
    attempts = 3

    while attempts > 0:
        input_username = input("Masukkan username: ")
        input_password = input("Masukkan password: ")

        if input_username == username and input_password == password:
            print("Login berhasil!")
            break
        else:
            attempts -= 1
            print("Username atau password salah. Sisa percobaan:", attempts)

    if attempts == 0:
        print("Anda telah melebihi batas percobaan. Akun terkunci.")
        quit()
login()


import openpyxl

# Fungsi untuk meminta input dari pengguna
def get_input(prompt):
    return input(prompt)

# Fungsi untuk menyimpan data pengajuan tandatangan dalam spreadsheet
def simpan_data(nama_pemohon, nama_kegiatan, deadline, deskripsi, tujuan, file_proposal, keterangan_pengajuan):
    wb = openpyxl.load_workbook('hasil_pengajuan_tandatangan.xlsx')
    sheet = wb.active

    # Mendapatkan jumlah baris yang sudah terisi
    row_count = sheet.max_row

    # Menyimpan data pengajuan tandatangan di baris berikutnya
    row_data = [nama_pemohon, nama_kegiatan, deadline, deskripsi, tujuan, file_proposal, keterangan_pengajuan]
    sheet.append(row_data)

    # Menyimpan file spreadsheet
    wb.save('hasil_pengajuan_tandatangan.xlsx')
    print('Data pengajuan tandatangan telah disimpan dalam file hasil_pengajuan_tandatangan.xlsx')
    import programTU

# Menampilkan data pengajuan tandatangan
def tampilkan_data():
    wb = openpyxl.load_workbook('hasil_pengajuan_tandatangan.xlsx')
    sheet = wb.active

    print("Data pengajuan tandatangan:")
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print("Nama Pemohon:", row[0])
        print("Nama Kegiatan:", row[1])
        print("Deadline Tandatangan:", row[2])
        print("Deskripsi Kegiatan:", row[3])
        print("Tujuan Tanda Tangan:", row[4])
        print("File Proposal Pengajuan:", row[5])
        print("Keterangan Pengajuan:", row[6])
        print("-----------------------------")

# Meminta input dari pengguna
nama_pemohon = get_input("Masukkan nama pemohon: ")
nama_kegiatan = get_input("Masukkan nama kegiatan: ")
deadline = get_input("Masukkan deadline tandatangan: ")
deskripsi = get_input("Masukkan deskripsi kegiatan: ")
tujuan = get_input("Masukkan nama dosen: ")
file_proposal = get_input("Masukkan nama file proposal pengajuan: ")
keterangan_pengajuan = get_input("ACC atau Belum: ")

# Memanggil fungsi untuk menyimpan data pengajuan tandatangan dalam spreadsheet
simpan_data(nama_pemohon, nama_kegiatan, deadline, deskripsi, tujuan, file_proposal, keterangan_pengajuan)


# Memanggil fungsi untuk menampilkan data pengajuan tandatangan
tampilkan_data()












