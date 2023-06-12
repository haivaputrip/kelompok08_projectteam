import openpyxl

def copy_row(source_file, source_sheet, destination_file, destination_sheet, row_number):
    try:
        # Load the source workbook
        source_workbook = openpyxl.load_workbook(source_file)
        
        # Load the destination workbook
        destination_workbook = openpyxl.load_workbook(destination_file)
        
        # Select the source sheet and destination sheet
        source_sheet = source_workbook[source_sheet]
        destination_sheet = destination_workbook[destination_sheet]
        
        # Copy the row from the source sheet to the destination sheet
        for column in range(1, 8):
            cell_value = source_sheet.cell(row=row_number, column=column).value
            destination_sheet.cell(row=row_number, column=column).value = cell_value
        
        # Save the changes to the destination file
        destination_workbook.save(destination_file)
        print("Row copied successfully.")
    
    except FileNotFoundError:
        print("File not found.")
    
    except Exception as e:
        print(f"An error occurred: {str(e)}")

# Example usage
source_file = "hasil_pengajuan_tandatangan.xlsx"
source_sheet = "Lembar1"
destination_file = "baris_FT.xlsx"
destination_sheet = "Lembar1"
sheet = destination_file.active
row_count = sheet.max_row
sheet.append(destination_file)
row_number = 8

copy_row(source_file, source_sheet, destination_file, destination_sheet, row_number)
