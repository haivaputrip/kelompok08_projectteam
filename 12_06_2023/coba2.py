import tkinter as tk
from tkinter import messagebox

def login():
    username = "admin"
    password = "password"
    attempts = 3

    def check_login():
        nonlocal attempts
        input_username = username_entry.get()
        input_password = password_entry.get()

        if input_username == username and input_password == password:
            messagebox.showinfo("Login", "Login berhasil!")
            login_window.destroy()
        else:
            attempts -= 1
            messagebox.showwarning("Login", f"Username atau password salah. Sisa percobaan: {attempts}")
            if attempts == 0:
                messagebox.showerror("Login", "Anda telah melebihi batas percobaan. Akun terkunci.")
                login_window.destroy()
                quit()
            

    # Membuat jendela login
    login_window = tk.Tk()
    login_window.title("Login")
    login_window.geometry("300x150")

    # Label dan input untuk username
    username_label = tk.Label(login_window, text="Username:")
    username_label.pack()
    username_entry = tk.Entry(login_window)
    username_entry.pack()

    # Label dan input untuk password
    password_label = tk.Label(login_window, text="Password:")
    password_label.pack()
    password_entry = tk.Entry(login_window, show="*")
    password_entry.pack()

    # Tombol login
    login_button = tk.Button(login_window, text="Login", command=check_login)
    login_button.pack()

    # Menjalankan jendela login
    login_window.mainloop()

# Memanggil fungsi login
login()

# Fungsi untuk menyimpan data pengajuan tandatangan dalam spreadsheet
import tkinter as tk
from tkinter import messagebox
import openpyxl

# Fungsi untuk menyimpan data pengajuan tandatangan dalam spreadsheet
def simpan_data():
    nama_pemohon = nama_pemohon_entry.get()
    nama_kegiatan = nama_kegiatan_entry.get()
    deadline = deadline_entry.get()
    deskripsi = deskripsi_entry.get()
    tujuan = tujuan_entry.get()
    file_proposal = file_proposal_entry.get()
    keterangan_pengajuan = keterangan_pengajuan_entry.get()

    wb = openpyxl.load_workbook('hasil_pengajuan_tandatangan.xlsx')
    sheet = wb.active

    # Mendapatkan jumlah baris yang sudah terisi
    row_count = sheet.max_row

    # Menyimpan data pengajuan tandatangan di baris berikutnya
    row_data = [nama_pemohon, nama_kegiatan, deadline, deskripsi, tujuan, file_proposal]
    sheet.append(row_data)

    # Menyimpan file spreadsheet
    wb.save('hasil_pengajuan_tandatangan.xlsx')
    messagebox.showinfo("Simpan Data", "Data pengajuan tandatangan telah disimpan dalam file hasil_pengajuan_tandatangan.xlsx")

# Menampilkan data pengajuan tandatangan
def tampilkan_data():
    wb = openpyxl.load_workbook('hasil_pengajuan_tandatangan.xlsx')
    sheet = wb.active

    data = "Data pengajuan tandatangan:\n\n"
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data += f"Nama Pemohon: {row[0]}\n"
        data += f"Nama Kegiatan: {row[1]}\n"
        data += f"Deadline Tandatangan: {row[2]}\n"
        data += f"Deskripsi Kegiatan: {row[3]}\n"
        data += f"Tujuan Tanda Tangan: {row[4]}\n"
        data += f"File Proposal Pengajuan: {row[5]}\n"
        data += f"keterangan Pengajuan: {row[6]}\n"
        data += "-----------------------------\n"

    messagebox.showinfo("Data Pengajuan Tandatangan", data)

# Membuat jendela pengajuan tandatangan
pengajuan_window = tk.Tk()
pengajuan_window.title("Pengajuan Tandatangan")
pengajuan_window.geometry("400x300")

# Label dan input untuk nama pemohon
nama_pemohon_label = tk.Label(pengajuan_window, text="Nama Pemohon:")
nama_pemohon_label.pack()
nama_pemohon_entry = tk.Entry(pengajuan_window)
nama_pemohon_entry.pack()

# Label dan input untuk nama kegiatan
nama_kegiatan_label = tk.Label(pengajuan_window, text="Nama Kegiatan:")
nama_kegiatan_label.pack()
nama_kegiatan_entry = tk.Entry(pengajuan_window)
nama_kegiatan_entry.pack()

# Label dan input untuk deadline tandatangan
deadline_label = tk.Label(pengajuan_window, text="Deadline Tandatangan:")
deadline_label.pack()
deadline_entry = tk.Entry(pengajuan_window)
deadline_entry.pack()

# Label dan input untuk deskripsi kegiatan
deskripsi_label = tk.Label(pengajuan_window, text="Deskripsi Kegiatan:")
deskripsi_label.pack()
deskripsi_entry = tk.Entry(pengajuan_window)
deskripsi_entry.pack()

# Label dan input untuk tujuan tanda tangan
tujuan_label = tk.Label(pengajuan_window, text="Tujuan Tanda Tangan:")
tujuan_label.pack()
tujuan_entry = tk.Entry(pengajuan_window)
tujuan_entry.pack()

# Label dan input untuk nama file proposal pengajuan
file_proposal_label = tk.Label(pengajuan_window, text="Nama File Proposal Pengajuan:")
file_proposal_label.pack()
file_proposal_entry = tk.Entry(pengajuan_window)
file_proposal_entry.pack()

# Label dan input untuk keterangan pengajuan
keterangan_pengajuan_label = tk.Label(pengajuan_window, text="ACC atau Belum:")
keterangan_pengajuan_label.pack()
keterangan_pengajuan_entry = tk.Entry(pengajuan_window)
keterangan_pengajuan_entry.pack()

# Tombol untuk menyimpan data pengajuan tandatangan
simpan_button = tk.Button(pengajuan_window, text="Simpan", command=simpan_data)
simpan_button.pack()

# Tombol untuk menampilkan data pengajuan tandatangan
tampilkan_button = tk.Button(pengajuan_window, text="Tampilkan Data", command=tampilkan_data)
tampilkan_button.pack()

# Menjalankan jendela pengajuan tandatangan
pengajuan_window.mainloop()