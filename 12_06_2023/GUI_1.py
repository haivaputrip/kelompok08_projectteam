import openpyxl
import tkinter as tk
from PIL import ImageTk, Image
from tkcalendar import DateEntry
from tkinter import messagebox
import pandas as pd
from tkinter import ttk
def get_input(prompt):
    return input(prompt)

def tampilkan_data():
    df = pd.read_csv('hasil_pengajuan_tandatangan.csv')

    root = tk.Tk()
    root.geometry("800x600")
    root.title("Data Pengajuan Tandatangan")

    treeview = ttk.Treeview(root, columns=list(df.columns), show='headings')

    for column in df.columns:
        treeview.heading(column, text=column)

    for row in df.itertuples(index=False):
        treeview.insert('', 'end', values=row)

    treeview.pack(expand=True, fill='both')
    root.mainloop()

def tombol8():
    mainform8 = tk.Tk()
    mainform8.geometry("1080x720")
    mainform8.resizable(False,False)

    Image1 = Image.open("logoft.PNG")
    Image1 = Image1.resize((434,200),Image.ANTIALIAS)
    photo = ImageTk.PhotoImage(Image1)
    gambar=tk.Label(mainform8,image=photo)
    gambar.place(x=270)
    gambar.configure(highlightthickness=0, borderwidth=0)
    gambar.place(x=330, y=0)

    kotak1 =tk.Frame(mainform8,width=300, height=30, bg= 'grey')
    kotak1.place(x=400,y=210)
    kotak2 = tk.Frame(mainform8,width=300, height=300, bg= 'white')
    kotak2.place(x=400,y=240)

    label3 = tk.Label(mainform8,text= "Masukkan Baris dan Nilai baru",font=("Times New Roman",12),bg="grey", fg= "white")
    label3.place(x=453,y=212)
    label4 = tk.Label(mainform8,text= "Masukkan Baris",font=("Times New Roman",13),bg= "white")
    label4.place(x=495,y=250)
    label5 = tk.Label(mainform8,text= "Masukkan Nilai Baru",font=("Times New Roman",13),bg= "white")
    label5.place(x=475,y=340)
    label6 = tk.Label(mainform8,text= "*Tambahkan 1 Angka pada Nomor Antrean",font=("Times New Roman",8),bg= "white",fg="red")
    label6.place(x=420,y=310)


    baris_var = tk.StringVar()
    nilai_baru_var = tk.StringVar()

    
        # Membaca file Excel

    def excel_to_csv(hasil_pengajuan_tandatangan, nama_file_csv):
        df = pd.read_excel(hasil_pengajuan_tandatangan)  # Membaca file Excel menjadi DataFrame
        df.to_csv(nama_file_csv, index=True)  # Menulis DataFrame ke file CSV

        print("File Excel berhasil diubah menjadi CSV.")

        # Contoh pemanggilan fungsi
    nama_file_excel = 'hasil_pengajuan_tandatangan.xlsx'  # Ganti dengan nama file Excel yang ingin diubah
    nama_file_csv = 'hasil_pengajuan_tandatangan.csv'     # Nama file CSV yang dihasilkan

    excel_to_csv(nama_file_excel, nama_file_csv)
        
       

    def buka_file_excel():
        nama_file = "hasil_pengajuan_tandatangan.xlsx"
        
        try:
            # Buka file Excel
            workbook = openpyxl.load_workbook(nama_file)
            
            # Mengambil sheet pertama
            sheet = workbook.active
            
            # Membaca data
            for row in sheet.iter_rows():
                for cell in row:
                    print(cell.value, end="\t")
                print()
            
            # Mengizinkan pengguna mengedit file

            baris = baris_var.get() # Mengambil nilai dari objek StringVar
            barisbaru = int(baris) + 1  
            
            kolom = "H"
            nilai_baru = nilai_baru_var.get()
                
            # Mengubah nilai sel yang diinginkan
            sheet[kolom + str(barisbaru)] = nilai_baru
            
            # Menyimpan perubahan ke file
            workbook.save(nama_file)
            print("Perubahan telah disimpan.")
        except FileNotFoundError:
            print("File tidak ditemukan.")
        
        except Exception as e:
            print(f"Terjadi kesalahan: {str(e)}")
    
    def notifkonfirm1():
        messagebox.showinfo("Confirm", "Data Berhasil Diubah!")
        mainform8.destroy()
        tombol2()

    ent1 = tk.Entry(mainform8, width=28, font=("Times New Roman", 14), textvariable=baris_var)  # Menghubungkan objek StringVar dengan Entry
    ent1.place(x=420, y=280, height=35)
    ent2 = tk.Entry(mainform8, width=28, font=("Times New Roman", 14), textvariable=nilai_baru_var)  # Menghubungkan objek StringVar dengan Entry
    ent2.place(x=420, y=370, height=35)
    
    button1 = tk.Button(mainform8, text="Submit", bg="blue", fg="white", font=("Times New Roman", 12), command=lambda:{buka_file_excel(),excel_to_csv(nama_file_excel, nama_file_csv),notifkonfirm1()})
    button1.config(width=10, height=1)
    button1.place(x=575, y=480)
    button2 = tk.Button(mainform8, text="Kembali", bg="red", fg="white", font=("Times New Roman", 12), command=lambda:{mainform8.destroy(),tombol2()})
    button2.config(width=10, height=1)
    button2.place(x=420, y=480)
    button3 = tk.Button(mainform8, text="Tampilkan Data", bg="grey", font=("Times New Roman", 12), command=lambda:{tampilkan_data()})
    button3.config(width=15, height=1)
    button3.place(x=475, y=430)
    mainform8.mainloop()

def tombol7():
    mainform7 = tk.Tk()
    mainform7.geometry("1080x720")
    mainform7.resizable(False,False)
    mainform7.configure(background="#395b7d")

    Image2 = Image.open("logoti2.PNG")
    Image2 = Image2.resize((275,275),Image.ANTIALIAS)
    photo1= ImageTk.PhotoImage(Image2)
    gambar1=tk.Label(mainform7,image=photo1,background="#395b7d")
    gambar1.place(x=210,y=-30)
    gambar1.configure(highlightthickness=0, borderwidth=0,)

    kotak1 =tk.Frame(mainform7,width=300, height=30, bg= 'grey')
    kotak1.place(x=400,y=200)
    kotak2 = tk.Frame(mainform7,width=300, height=300, bg= 'white')
    kotak2.place(x=400,y=230)

    label1= tk.Label(mainform7,text= "Teknik Industri",font=("Times New Roman",44),bg="#395b7d", fg= "white")
    label1.place(x=435,y=40)
    label2= tk.Label(mainform7,text= "Universitas Sebelas Maret",font=("Times New Roman",30),bg="#395b7d", fg= "white")
    label2.place(x=435,y=110)
    label3 = tk.Label(mainform7,text= "Masukkan Baris dan Nilai baru",font=("Times New Roman",12),bg="grey", fg= "white")
    label3.place(x=453,y=203)
    label4 = tk.Label(mainform7,text= "Masukkan Baris",font=("Times New Roman",13),bg= "white")
    label4.place(x=495,y=250)
    label5 = tk.Label(mainform7,text= "Masukkan Nilai Baru",font=("Times New Roman",13),bg= "white")
    label5.place(x=475,y=340)
    label6 = tk.Label(mainform7,text= "*Tambahkan 1 Angka pada Nomor Antrean",font=("Times New Roman",8),bg= "white",fg="red")
    label6.place(x=420,y=310)

    baris_var = tk.StringVar()
    nilai_baru_var = tk.StringVar()

        # Membaca file Excel

    def excel_to_csv(hasil_pengajuan_tandatangan, nama_file_csv):
        df = pd.read_excel(hasil_pengajuan_tandatangan)  # Membaca file Excel menjadi DataFrame
        df.to_csv(nama_file_csv, index=True)  # Menulis DataFrame ke file CSV

        print("File Excel berhasil diubah menjadi CSV.")

        # Contoh pemanggilan fungsi
    nama_file_excel = 'hasil_pengajuan_tandatangan.xlsx'  # Ganti dengan nama file Excel yang ingin diubah
    nama_file_csv = 'hasil_pengajuan_tandatangan.csv'     # Nama file CSV yang dihasilkan

    excel_to_csv(nama_file_excel, nama_file_csv)

    def buka_file_excel():
        nama_file = "hasil_pengajuan_tandatangan.xlsx"
        
        try:
            # Buka file Excel
            workbook = openpyxl.load_workbook(nama_file)
            
            # Mengambil sheet pertama
            sheet = workbook.active
            
            # Membaca data
            for row in sheet.iter_rows():
                for cell in row:
                    print(cell.value, end="\t")
                print()
            
            # Mengizinkan pengguna mengedit file

            baris = baris_var.get() # Mengambil nilai dari objek StringVar
            barisbaru = int(baris) + 1

            # if int(baris) == 0:
            #     quit()    
            
            kolom = "G"
            nilai_baru = nilai_baru_var.get()
                
            # Mengubah nilai sel yang diinginkan
            sheet[kolom + str(barisbaru)] = nilai_baru
            
            # Menyimpan perubahan ke file
            workbook.save(nama_file)
            print("Perubahan telah disimpan.")
        except FileNotFoundError:
            print("File tidak ditemukan.")
        
        except Exception as e:
            print(f"Terjadi kesalahan: {str(e)}")
    def notifkonfirm():
        messagebox.showinfo("Confirm", "Data Berhasil Diubah!")
        mainform7.destroy()
        tombol4()

    ent1 = tk.Entry(mainform7, width=28, font=("Times New Roman", 14), textvariable=baris_var)  # Menghubungkan objek StringVar dengan Entry
    ent1.place(x=420, y=280, height=35)
    ent2 = tk.Entry(mainform7, width=28, font=("Times New Roman", 14), textvariable=nilai_baru_var)  # Menghubungkan objek StringVar dengan Entry
    ent2.place(x=420, y=370, height=35)
    
    button1 = tk.Button(mainform7, text="Submit", bg="blue", fg="white", font=("Times New Roman", 12), command=lambda:{buka_file_excel(),excel_to_csv(nama_file_excel, nama_file_csv),notifkonfirm()})
    button1.config(width=10, height=1)
    button1.place(x=575, y=480)
    button2 = tk.Button(mainform7, text="Kembali", bg="red", fg="white", font=("Times New Roman", 12),command=lambda:{mainform7.destroy(),tombol4()})
    button2.config(width=10, height=1)
    button2.place(x=420, y=480)
    button3 = tk.Button(mainform7, text="Tampilkan Data", bg="grey", font=("Times New Roman", 12), command=lambda:{tampilkan_data()})
    button3.config(width=15, height=1)
    button3.place(x=475, y=430)
    mainform7.mainloop()


def simpan_data(nama_pemohon, nama_kegiatan, deadline, deskripsi, tujuan, file_proposal, keterangan_pengajuan,keterangan_fakultas):
    wb = openpyxl.load_workbook('hasil_pengajuan_tandatangan.xlsx')
    sheet = wb.active

    row_count = sheet.max_row
    row_data = [nama_pemohon, nama_kegiatan, deadline, deskripsi, tujuan, file_proposal, keterangan_pengajuan,keterangan_fakultas]
    sheet.append(row_data)

    wb.save('hasil_pengajuan_tandatangan.xlsx')
    print('Data pengajuan tandatangan telah disimpan dalam file hasil_pengajuan_tandatangan.xlsx')

def tombol6():
    mainform6 = tk.Tk()
    mainform6.geometry("1080x720")
    mainform6.resizable(False,False)
    mainform6.configure(background="#395b7d")

    Image2 = Image.open("logoti2.PNG")
    Image2 = Image2.resize((275,275),Image.ANTIALIAS)
    photo1= ImageTk.PhotoImage(Image2)
    gambar1=tk.Label(mainform6,image=photo1,background="#395b7d")
    gambar1.place(x=210,y=-30)
    gambar1.configure(highlightthickness=0, borderwidth=0,)

    kotak1 =tk.Frame(mainform6,width=905, height=400, bg= 'white')
    kotak1.place(x=90,y=210)

    label1= tk.Label(mainform6,text= "Teknik Industri",font=("Times New Roman",44),bg="#395b7d", fg= "white")
    label1.place(x=435,y=40)
    label2= tk.Label(mainform6,text= "Universitas Sebelas Maret",font=("Times New Roman",30),bg="#395b7d", fg= "white")
    label2.place(x=435,y=110)
    label3= tk.Label(mainform6,text= "Masukkan Data Pemohon",font=("Times New Roman",14),bg="grey", fg= "white",width=90)
    label3.place(x=90,y=210)


    ent1 =tk.Entry(mainform6,width=35, font=("Times New Roman", 12))
    ent1.place(x=170,y=290)
    label4=tk.Label(mainform6,text= "Nama Pemohon",font=("Times New Roman",12), bg="white")
    label4.place(x=170,y=265)

    ent2 =tk.Entry(mainform6,width=35, font=("Times New Roman", 12))
    ent2.place(x=170,y=370)
    label5=tk.Label(mainform6,text= "Nama Kegiatan",font=("Times New Roman",12), bg="white")
    label5.place(x=170,y=345)

    ent3 =DateEntry(mainform6,width=35, font=("Times New Roman", 12))
    ent3.place(x=170,y=450)
    label6=tk.Label(mainform6,text= "Deadline Tanda Tangan",font=("Times New Roman",12), bg="white")
    label6.place(x=170,y=425)

    ent4 =tk.Entry(mainform6,width=35, font=("Times New Roman", 12))
    ent4.place(x=600,y=290)
    label7=tk.Label(mainform6,text= "Nama Dosen Tujuan",font=("Times New Roman",12), bg="white")
    label7.place(x=600,y=265)

    ent5 =tk.Entry(mainform6,width=35, font=("Times New Roman", 12))
    ent5.place(x=600,y=370)
    label8=tk.Label(mainform6,text= "Proposal Lanjut FT(Ya/Tidak)",font=("Times New Roman",12), bg="white")
    label8.place(x=600,y=345)

    ent6 =tk.Entry(mainform6,width=35, font=("Times New Roman", 12))
    ent6.place(x=600,y=450)
    label9=tk.Label(mainform6,text= "Keterangan",font=("Times New Roman",12), bg="white")
    label9.place(x=600,y=425)

    ent7 =tk.Entry(mainform6, font=("Times New Roman", 12))
    ent7.place(x=350,y=510,width=400, height=30)
    label10=tk.Label(mainform6,text= "Deskripsi Kegiatan",font=("Times New Roman",12), bg="white")
    label10.place(x=350,y=485)

    def submit_data():
        nama_pemohon = ent1.get()
        nama_kegiatan = ent2.get()
        deadline = ent3.get()
        tujuan = ent4.get()
        file_proposal = ent5.get()
        keterangan_pengajuan = ent6.get()
        deskripsi = ent7.get()
        keterangan_fakultas = ""

        simpan_data(nama_pemohon, nama_kegiatan, deadline, deskripsi, tujuan, file_proposal, keterangan_pengajuan,keterangan_fakultas)
    def exceltocsv():
        data_frame = pd.read_excel('hasil_pengajuan_tandatangan.xlsx')

        # Simpan ke file CSV
        data_frame.to_csv('hasil_pengajuan_tandatangan.csv', index=True)

    def notifantre():
        messagebox.showinfo("Submit", "Data Berhasil ditambahkan!")
        mainform6.destroy()
        tombol4()
    
    button1 = tk.Button(mainform6, text="Submit", bg="blue", fg="white", font=("Times New Roman", 12), command=lambda:{submit_data(),exceltocsv(),notifantre()})
    button1.config(width=10, height=1)
    button1.place(x=870, y=560)
    button2 = tk.Button(mainform6, text="Kembali", bg="red", fg="white", font=("Times New Roman", 12), command=lambda:{mainform6.destroy(),tombol4()})
    button2.config(width=10, height=1)
    button2.place(x=120, y=560)

    mainform6.mainloop()

def tombol5():
    def login1():
        username = "fakultas"
        password = "teknik"
        attempts = 3

        def check_login1():
            nonlocal attempts
            input_username = ent3.get()
            input_password = ent4.get()

            if input_username == username and input_password == password:
                messagebox.showinfo("Login", "Login berhasil!")
                mainform5.destroy()
                tombol8()
            else:
                attempts -= 1
                messagebox.showwarning("Login", f"Username atau password salah. Sisa percobaan: {attempts}")
                if attempts == 0:
                    messagebox.showerror("Login", "Anda telah melebihi batas percobaan. Akun terkunci.")
                    mainform5.destroy()
                    quit()
        mainform5 = tk.Tk()
        mainform5.geometry("1080x720")
        mainform5.resizable(False,False)
    

        Image1 = Image.open("logoft.PNG")
        Image1 = Image1.resize((434,200),Image.ANTIALIAS)
        photo = ImageTk.PhotoImage(Image1)
        gambar=tk.Label(mainform5,image=photo)
        gambar.place(x=270)
        gambar.configure(highlightthickness=0, borderwidth=0)
        gambar.place(x=325, y=0)


        kotak1 =tk.Frame(mainform5,width=300, height=30, bg= 'grey')
        kotak1.place(x=400,y=210)
        kotak2 = tk.Frame(mainform5,width=300, height=300, bg= 'white')
        kotak2.place(x=400,y=240)

        label2 = tk.Label(mainform5,text= "Masukkan Username dan Password Anda",font=("Times New Roman",12),bg="grey", fg= "white")
        label2.place(x=420,y=213)
        label3 = tk.Label(mainform5,text= "Username",font=("Times New Roman",12),bg="white")
        label3.place(x=515,y=270)
        label4 = tk.Label(mainform5,text= "Password",font=("Times New Roman",12),bg="white")
        label4.place(x=515,y=330)

        ent3 = tk.Entry(mainform5,width=20, font=("Times New Roman", 15) )
        ent3.place(x=450,y=300)
        ent4 = tk.Entry(mainform5,width=20, font=("Times New Roman", 15), show="*" )
        ent4.place(x=450,y=360)

        button1= tk.Button(mainform5,text="Masuk",bg="blue",fg="white", font=("Times New Roman",12), command= lambda:{check_login1()})
        button1.config(width=10,height=1)
        button1.place(x=497,y=410)
        button2= tk.Button(mainform5,text="Kembali",bg="red",fg="white", font=("Times New Roman",12),command=lambda:{mainform5.destroy(),tombol2()})
        button2.config(width=6,height=1)
        button2.place(x=515,y=450)

        mainform5.mainloop()
    login1()

def tombol4():
    mainform4 = tk.Tk()
    mainform4.geometry("1080x720")
    mainform4.resizable(False,False)
    mainform4.configure(background="#395b7d")

    Image2 = Image.open("logoti2.PNG")
    Image2 = Image2.resize((275,275),Image.ANTIALIAS)
    photo1= ImageTk.PhotoImage(Image2)
    gambar1=tk.Label(mainform4,image=photo1,background="#395b7d")
    gambar1.place(x=210,y=-30)
    gambar1.configure(highlightthickness=0, borderwidth=0,)

    label0= tk.Label(mainform4,text= "Teknik Industri",font=("Times New Roman",44),bg="#395b7d", fg= "white")
    label0.place(x=435,y=40)
    label1= tk.Label(mainform4,text= "Universitas Sebelas Maret",font=("Times New Roman",30),bg="#395b7d", fg= "white")
    label1.place(x=435,y=110)

    kotak1 =tk.Frame(mainform4,width=300, height=30, bg= 'grey')
    kotak1.place(x=400,y=250)
    label2 = tk.Label(mainform4,text= "Pilih Keperluan",font=("Times New Roman",12),bg="grey", fg= "white")
    label2.place(x=500,y=252)
    kotak2 = tk.Frame(mainform4,width=300, height=300, bg= 'white')
    kotak2.place(x=400,y=280)

    button2 = tk.Button(mainform4,text="Masukkan Antrian Prodi",command= lambda:{mainform4.destroy(),tombol6()})
    button2.config(width=30, height=3)
    button2.place(x=443,y=320)
    button3 = tk.Button(mainform4,text="Konfirmasi Prodi",command=lambda:{mainform4.destroy(),tombol7()})
    button3.config(width=30, height=3)
    button3.place(x=443,y=400)
  
    button5 = tk.Button(mainform4,text="Kembali", bg="red",fg="white",command=lambda:{mainform4.destroy(),tombol2()})
    button5.config(width=30, height=2)
    button5.place(x=443,y=500)

    mainform4.mainloop()

def tombol3():
    def login():
        username = "admin"
        password = "password"
        attempts = 3

        def check_login():
            nonlocal attempts
            input_username = ent1.get()
            input_password = ent2.get()

            if input_username == username and input_password == password:
                messagebox.showinfo("Login", "Login berhasil!")
                mainform3.destroy()
                tombol4()
            else:
                attempts -= 1
                messagebox.showwarning("Login", f"Username atau password salah. Sisa percobaan: {attempts}")
                if attempts == 0:
                    messagebox.showerror("Login", "Anda telah melebihi batas percobaan. Akun terkunci.")
                    mainform3.destroy()
                    quit()

        mainform3 = tk.Tk()
        mainform3.geometry("1080x720")
        mainform3.resizable(False,False)
        mainform3.configure(background="#395b7d")

        Image2 = Image.open("logoti2.PNG")
        Image2 = Image2.resize((275,275),Image.ANTIALIAS)
        photo1= ImageTk.PhotoImage(Image2)
        gambar1=tk.Label(mainform3,image=photo1,background="#395b7d")
        gambar1.place(x=210,y=-30)
        gambar1.configure(highlightthickness=0, borderwidth=0,)


        kotak1 =tk.Frame(mainform3,width=300, height=30, bg= 'grey')
        kotak1.place(x=400,y=200)
        kotak2 = tk.Frame(mainform3,width=300, height=300, bg= 'white')
        kotak2.place(x=400,y=230)

        label0= tk.Label(mainform3,text= "Teknik Industri",font=("Times New Roman",44),bg="#395b7d", fg= "white")
        label0.place(x=435,y=40)
        label1= tk.Label(mainform3,text= "Universitas Sebelas Maret",font=("Times New Roman",30),bg="#395b7d", fg= "white")
        label1.place(x=435,y=110)
        label2 = tk.Label(mainform3,text= "Masukkan Username dan Password Anda",font=("Times New Roman",12),bg="grey", fg= "white")
        label2.place(x=420,y=203)
        label3 = tk.Label(mainform3,text= "Username",font=("Times New Roman",12),bg="white")
        label3.place(x=515,y=260)
        label4 = tk.Label(mainform3,text= "Password",font=("Times New Roman",12),bg="white")
        label4.place(x=515,y=320)

        ent1 = tk.Entry(mainform3,width=20, font=("Times New Roman", 15))
        ent1.place(x=450,y=290)
        ent2 = tk.Entry(mainform3,width=20, font=("Times New Roman", 15), show="*" )
        ent2.place(x=450,y=350)


        button1= tk.Button(mainform3,text="Masuk",bg="blue",fg="white", font=("Times New Roman",12),command= lambda:check_login())
        button1.config(width=10,height=1)
        button1.place(x=497,y=400)
        button2= tk.Button(mainform3,text="Kembali",bg="red",fg="white", font=("Times New Roman",12),command=lambda:{mainform3.destroy(),tombol2()})
        button2.config(width=6,height=1)
        button2.place(x=515,y=440)

        mainform3.mainloop()
    login()

    
def tombol2():
    
    mainform2 = tk.Tk()
    mainform2.geometry("1080x720")
    mainform2.resizable(False,False)
    
    Image1 = Image.open("FT1.JPG")
    Image1 = Image1.resize((1080,720),Image.ANTIALIAS)
    photo = ImageTk.PhotoImage(Image1)
    gambar=tk.Label(mainform2,image=photo)
    gambar.place(x=0,y=0)

    label1 = tk.Label(mainform2,text="Selamat Datang",font=("Algerian",18))
    label1.place(x=460,y=120)
    label2 = tk.Label(mainform2,text="Di",font=("Algerian",15))
    label2.place(x=555,y=160)
    label3 = tk.Label(mainform2,text="Tata Usaha Teknik Industri",font=("Algerian",28))
    label3.place(x=300,y=200)

    button2 = tk.Button(mainform2,text="Pihak Program Studi",command= lambda:{mainform2.destroy(), tombol3()})
    button2.config(width=30,height=1)
    button2.place(x=465,y=340)
    button3 = tk.Button(mainform2,text="Pihak Fakultas", command= lambda:{mainform2.destroy(),tombol5()})
    button3.config(width=20,height=1)
    button3.place(x=500,y=290)
    button4 = tk.Button(mainform2,text="Lihat Antrean",background="blue",fg="white", command= lambda:{mainform2.destroy(),tampilkan_data()})
    button4.config(width=20,height=2)
    button4.place(x=500,y=390)

    mainform2.mainloop()    


mainform1 = tk.Tk()
mainform1.geometry("1080x720")
mainform1.resizable(False,False)

Image1 = Image.open("FT1.JPG")
Image1 = Image1.resize((1080,720),Image.ANTIALIAS)
photo = ImageTk.PhotoImage(Image1)

gambar=tk.Label(mainform1,image=photo)
gambar.place(x=0,y=0)

label1 = tk.Label(mainform1,text="Selamat Datang",font=("Algerian",18))
label1.place(x=460,y=120)
label2 = tk.Label(mainform1,text="Di",font=("Algerian",15))
label2.place(x=555,y=160)
label3 = tk.Label(mainform1,text="ADMINISTRASI FAKULTAS TEKNIK",font=("Algerian",28))
label3.place(x=270,y=200)

button1 = tk.Button(mainform1,text="Mulai",command=lambda:{mainform1.destroy(),tombol2()})
button1.config(width=10,height=2)
button1.place(x=515,y=320)

mainform1.mainloop()